"""
Supplier Listings API - Manage product catalog

Dual pricing: svaki listing ima price_rsd i price_eur.
Dobavljac unosi jednu cenu, sistem auto-preracunava drugu po kursu.
Ako su obe unete, EUR ima prioritet.
"""
from flask import Blueprint, request, g, send_file
from app.extensions import db
from app.models import SupplierListing, Supplier
from .auth import supplier_jwt_required
from app.utils.file_security import validate_upload
from app.constants.brands import get_brand_list, validate_brand
from pydantic import BaseModel, Field, model_validator
from typing import Optional
from datetime import datetime
from decimal import Decimal
from io import BytesIO
import logging

bp = Blueprint('supplier_listings', __name__, url_prefix='/listings')
logger = logging.getLogger(__name__)

# Max limits za Excel import
MAX_IMPORT_FILE_SIZE_MB = 5
MAX_IMPORT_ROWS = 5000


# ============== Price Helper ==============

def calculate_dual_prices(price_rsd=None, price_eur=None, eur_rate=117.5):
    """
    Racuna obe cene. EUR ima prioritet.

    Pravila:
    1. Ako je data SAMO price_eur -> price_rsd = price_eur * eur_rate
    2. Ako je data SAMO price_rsd -> price_eur = price_rsd / eur_rate
    3. Ako su date OBE -> EUR prioritet, price_rsd = price_eur * eur_rate
    4. Ako nije data NI JEDNA -> ValueError
    """
    rate = Decimal(str(eur_rate))

    if price_eur is not None:
        eur = Decimal(str(price_eur))
        return {
            'price_eur': eur.quantize(Decimal('0.01')),
            'price_rsd': (eur * rate).quantize(Decimal('0.01'))
        }
    elif price_rsd is not None:
        rsd = Decimal(str(price_rsd))
        return {
            'price_rsd': rsd.quantize(Decimal('0.01')),
            'price_eur': (rsd / rate).quantize(Decimal('0.01'))
        }
    else:
        raise ValueError('Mora biti unesena bar jedna cena (RSD ili EUR)')


# ============== Excel Column Mapping ==============

COLUMN_MAP = {
    # Naziv artikla
    'naziv': 'name',
    'ime': 'name',
    'name': 'name',
    'artikal': 'name',
    'opis_artikla': 'name',

    # Sifra
    'sifra': 'part_number',
    'šifra': 'part_number',
    'part_number': 'part_number',
    'code': 'part_number',
    'kod': 'part_number',
    'sku': 'part_number',
    'kataloški_broj': 'part_number',
    'kataloski_broj': 'part_number',
    'ref': 'part_number',

    # Brend
    'brend': 'brand',
    'brand': 'brand',
    'proizvodjac': 'brand',
    'proizvođač': 'brand',
    'marka': 'brand',

    # Kategorija
    'kategorija': 'part_category',
    'category': 'part_category',
    'part_category': 'part_category',
    'tip': 'part_category',
    'vrsta': 'part_category',

    # Cena RSD
    'cena_rsd': 'price_rsd',
    'price_rsd': 'price_rsd',
    'cena': 'price_rsd',
    'cena_rsd': 'price_rsd',
    'rsd': 'price_rsd',
    'cena_(rsd)': 'price_rsd',

    # Cena EUR
    'cena_eur': 'price_eur',
    'price_eur': 'price_eur',
    'eur': 'price_eur',
    'cena_(eur)': 'price_eur',
    'price': 'price_eur',  # default "price" without currency = EUR (common for suppliers)
    'cijena': 'price_eur',

    # Kolicina
    'kolicina': 'stock_quantity',
    'količina': 'stock_quantity',
    'stock': 'stock_quantity',
    'stanje': 'stock_quantity',
    'stock_quantity': 'stock_quantity',
    'qty': 'stock_quantity',
    'kom': 'stock_quantity',
    'na_stanju': 'stock_quantity',

    # Kvalitet
    'kvalitet': 'quality_grade',
    'quality': 'quality_grade',
    'quality_grade': 'quality_grade',
    'klasa': 'quality_grade',
    'grade': 'quality_grade',
    'tip_kvaliteta': 'quality_grade',

    # Model
    'model': 'model_compatibility',
    'modeli': 'model_compatibility',
    'model_compatibility': 'model_compatibility',
    'kompatibilnost': 'model_compatibility',
    'za_model': 'model_compatibility',
    'compatibility': 'model_compatibility',

    # Rok isporuke
    'rok_isporuke': 'delivery_days',
    'delivery_days': 'delivery_days',
    'isporuka': 'delivery_days',
    'rok': 'delivery_days',
    'delivery': 'delivery_days',

    # Opis
    'opis': 'description',
    'description': 'description',
    'napomena': 'description',
    'note': 'description',
    'notes': 'description',
}


# ============== Pydantic Schemas ==============

class ListingCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=200)
    brand: Optional[str] = Field(None, max_length=50)
    model_compatibility: Optional[str] = None
    part_category: Optional[str] = Field(None, max_length=50)
    part_number: Optional[str] = Field(None, max_length=50)
    description: Optional[str] = None
    is_original: Optional[bool] = False
    quality_grade: Optional[str] = Field(None, max_length=20)
    price_rsd: Optional[float] = Field(None, gt=0)
    price_eur: Optional[float] = Field(None, gt=0)
    min_order_qty: Optional[int] = Field(1, ge=1)
    stock_quantity: Optional[int] = Field(0, ge=0)
    stock_status: Optional[str] = Field(None, max_length=20)
    delivery_days: Optional[int] = Field(None, ge=0)

    @model_validator(mode='after')
    def at_least_one_price(self):
        if self.price_rsd is None and self.price_eur is None:
            raise ValueError('Mora biti unesena bar jedna cena (RSD ili EUR)')
        return self


class ListingUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=2, max_length=200)
    brand: Optional[str] = Field(None, max_length=50)
    model_compatibility: Optional[str] = None
    part_category: Optional[str] = Field(None, max_length=50)
    part_number: Optional[str] = Field(None, max_length=50)
    description: Optional[str] = None
    is_original: Optional[bool] = None
    quality_grade: Optional[str] = Field(None, max_length=20)
    price_rsd: Optional[float] = Field(None, gt=0)
    price_eur: Optional[float] = Field(None, gt=0)
    min_order_qty: Optional[int] = Field(None, ge=1)
    stock_quantity: Optional[int] = Field(None, ge=0)
    stock_status: Optional[str] = Field(None, max_length=20)
    delivery_days: Optional[int] = Field(None, ge=0)
    is_active: Optional[bool] = None


class BulkStockUpdate(BaseModel):
    listing_id: int
    stock_quantity: int = Field(..., ge=0)


# ============== Helper: Build listing response ==============

def listing_to_response(l):
    """Standardizovani response za listing."""
    return {
        'id': l.id,
        'name': l.name,
        'brand': l.brand,
        'model_compatibility': l.model_compatibility,
        'part_category': l.part_category,
        'part_number': l.part_number,
        'description': l.description,
        'is_original': l.is_original,
        'quality_grade': l.quality_grade,
        'price_rsd': float(l.price_rsd) if l.price_rsd else None,
        'price_eur': float(l.price_eur) if l.price_eur else None,
        'min_order_qty': l.min_order_qty or 1,
        'stock_quantity': l.stock_quantity,
        'stock_status': l.stock_status,
        'delivery_days': l.delivery_days,
        'is_active': l.is_active,
        'created_at': l.created_at.isoformat() if l.created_at else None,
        'updated_at': l.updated_at.isoformat() if l.updated_at else None
    }


def get_supplier_eur_rate():
    """Vrati EUR rate za trenutnog dobavljaca."""
    supplier = Supplier.query.get(g.supplier_id)
    if supplier and supplier.eur_rate:
        return float(supplier.eur_rate)
    return 117.5


# ============== Routes ==============

@bp.route('', methods=['GET'])
@supplier_jwt_required
def list_listings():
    """List all listings for current supplier"""
    include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
    category = request.args.get('category')
    brand = request.args.get('brand')
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = min(request.args.get('per_page', 50, type=int), 200)

    query = SupplierListing.query.filter_by(supplier_id=g.supplier_id)

    if not include_inactive:
        query = query.filter_by(is_active=True)

    if category:
        query = query.filter_by(part_category=category)

    if brand:
        query = query.filter(SupplierListing.brand.ilike(f'%{brand}%'))

    if q:
        query = query.filter(
            db.or_(
                SupplierListing.name.ilike(f'%{q}%'),
                SupplierListing.part_number.ilike(f'%{q}%'),
                SupplierListing.brand.ilike(f'%{q}%')
            )
        )

    query = query.order_by(SupplierListing.name)
    total = query.count()
    listings = query.offset((page - 1) * per_page).limit(per_page).all()

    return {
        'listings': [listing_to_response(l) for l in listings],
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
        'eur_rate': get_supplier_eur_rate()
    }


@bp.route('/<int:listing_id>', methods=['GET'])
@supplier_jwt_required
def get_listing(listing_id):
    """Get single listing details"""
    listing = SupplierListing.query.filter_by(
        id=listing_id,
        supplier_id=g.supplier_id
    ).first()

    if not listing:
        return {'error': 'Listing not found'}, 404

    return listing_to_response(listing)


@bp.route('', methods=['POST'])
@supplier_jwt_required
def create_listing():
    """Create new listing with dual pricing"""
    try:
        data = ListingCreate(**request.json)
    except Exception as e:
        return {'error': str(e)}, 400

    eur_rate = get_supplier_eur_rate()
    prices = calculate_dual_prices(
        price_rsd=data.price_rsd,
        price_eur=data.price_eur,
        eur_rate=eur_rate
    )

    listing = SupplierListing(
        supplier_id=g.supplier_id,
        name=data.name,
        brand=data.brand,
        model_compatibility=data.model_compatibility,
        part_category=data.part_category,
        part_number=data.part_number,
        description=data.description,
        is_original=data.is_original,
        quality_grade=data.quality_grade,
        price_rsd=prices['price_rsd'],
        price_eur=prices['price_eur'],
        price=prices['price_rsd'],  # Legacy backward compat
        currency='RSD',
        min_order_qty=data.min_order_qty or 1,
        stock_quantity=data.stock_quantity or 0,
        stock_status=data.stock_status,
        delivery_days=data.delivery_days,
        is_active=True
    )

    db.session.add(listing)
    db.session.commit()

    return {
        'message': 'Listing created',
        'listing_id': listing.id
    }, 201


@bp.route('/<int:listing_id>', methods=['PUT'])
@supplier_jwt_required
def update_listing(listing_id):
    """Update listing with dual pricing"""
    listing = SupplierListing.query.filter_by(
        id=listing_id,
        supplier_id=g.supplier_id
    ).first()

    if not listing:
        return {'error': 'Listing not found'}, 404

    try:
        data = ListingUpdate(**request.json)
    except Exception as e:
        return {'error': str(e)}, 400

    if data.name is not None:
        listing.name = data.name
    if data.brand is not None:
        listing.brand = data.brand
    if data.model_compatibility is not None:
        listing.model_compatibility = data.model_compatibility
    if data.part_category is not None:
        listing.part_category = data.part_category
    if data.part_number is not None:
        listing.part_number = data.part_number
    if data.description is not None:
        listing.description = data.description
    if data.is_original is not None:
        listing.is_original = data.is_original
    if data.quality_grade is not None:
        listing.quality_grade = data.quality_grade

    # Dual pricing update
    if data.price_rsd is not None or data.price_eur is not None:
        eur_rate = get_supplier_eur_rate()
        prices = calculate_dual_prices(
            price_rsd=data.price_rsd,
            price_eur=data.price_eur,
            eur_rate=eur_rate
        )
        listing.price_rsd = prices['price_rsd']
        listing.price_eur = prices['price_eur']
        listing.price = prices['price_rsd']  # Legacy

    if data.min_order_qty is not None:
        listing.min_order_qty = data.min_order_qty
    if data.stock_quantity is not None:
        listing.stock_quantity = data.stock_quantity
    if data.stock_status is not None:
        listing.stock_status = data.stock_status
    if data.delivery_days is not None:
        listing.delivery_days = data.delivery_days
    if data.is_active is not None:
        listing.is_active = data.is_active

    listing.updated_at = datetime.utcnow()
    db.session.commit()

    return {'message': 'Listing updated', 'listing_id': listing.id}


@bp.route('/<int:listing_id>', methods=['DELETE'])
@supplier_jwt_required
def delete_listing(listing_id):
    """Delete listing (soft delete)"""
    listing = SupplierListing.query.filter_by(
        id=listing_id,
        supplier_id=g.supplier_id
    ).first()

    if not listing:
        return {'error': 'Listing not found'}, 404

    listing.is_active = False
    listing.updated_at = datetime.utcnow()
    db.session.commit()

    return {'message': 'Listing deleted'}


# ============== Toggle ==============

@bp.route('/<int:listing_id>/toggle', methods=['POST'])
@supplier_jwt_required
def toggle_listing(listing_id):
    """Toggle listing active/inactive"""
    listing = SupplierListing.query.filter_by(
        id=listing_id,
        supplier_id=g.supplier_id
    ).first()

    if not listing:
        return {'error': 'Listing not found'}, 404

    listing.is_active = not listing.is_active
    listing.updated_at = datetime.utcnow()
    db.session.commit()

    return {
        'success': True,
        'data': {
            'listing_id': listing.id,
            'is_active': listing.is_active
        }
    }


# ============== Bulk Stock ==============

@bp.route('/bulk-stock', methods=['PUT'])
@supplier_jwt_required
def bulk_update_stock():
    """Bulk update stock quantities"""
    data = request.json or {}
    updates = data.get('updates', [])

    if not updates:
        return {'error': 'No updates provided'}, 400

    updated_count = 0
    errors = []

    for update in updates:
        try:
            listing_id = update.get('listing_id')
            stock_quantity = update.get('stock_quantity')

            if listing_id is None or stock_quantity is None:
                errors.append(f'Invalid update: {update}')
                continue

            listing = SupplierListing.query.filter_by(
                id=listing_id,
                supplier_id=g.supplier_id
            ).first()

            if not listing:
                errors.append(f'Listing {listing_id} not found')
                continue

            listing.stock_quantity = max(0, int(stock_quantity))
            listing.updated_at = datetime.utcnow()
            updated_count += 1

        except Exception as e:
            errors.append(f'Error updating {listing_id}: {str(e)}')

    db.session.commit()

    return {
        'message': f'Updated {updated_count} listings',
        'updated_count': updated_count,
        'errors': errors if errors else None
    }


# ============== Bulk Price Change ==============

@bp.route('/bulk-price', methods=['PUT'])
@supplier_jwt_required
def bulk_change_price():
    """Change price by percentage for multiple listings"""
    data = request.json or {}
    listing_ids = data.get('listing_ids', [])
    percent_change = data.get('percent_change')

    if not listing_ids:
        return {'error': 'Nema selektovanih artikala'}, 400

    if percent_change is None:
        return {'error': 'percent_change je obavezan'}, 400

    try:
        pct = Decimal(str(percent_change))
    except Exception:
        return {'error': 'Nevalidan procenat'}, 400

    if pct == 0:
        return {'error': 'Procenat ne moze biti 0'}, 400

    multiplier = (Decimal('100') + pct) / Decimal('100')
    updated = 0

    listings = SupplierListing.query.filter(
        SupplierListing.id.in_(listing_ids),
        SupplierListing.supplier_id == g.supplier_id
    ).all()

    for listing in listings:
        if listing.price_eur:
            listing.price_eur = (listing.price_eur * multiplier).quantize(Decimal('0.01'))
        if listing.price_rsd:
            listing.price_rsd = (listing.price_rsd * multiplier).quantize(Decimal('0.01'))
        if listing.price:
            listing.price = listing.price_rsd  # Legacy sync
        listing.updated_at = datetime.utcnow()
        updated += 1

    db.session.commit()

    return {
        'success': True,
        'message': f'Azurirano {updated} artikala ({percent_change:+g}%)',
        'updated': updated
    }


# ============== Bulk Toggle ==============

@bp.route('/bulk-toggle', methods=['PUT'])
@supplier_jwt_required
def bulk_toggle():
    """Activate or deactivate multiple listings at once"""
    data = request.json or {}
    listing_ids = data.get('listing_ids', [])
    is_active = data.get('is_active')

    if not listing_ids:
        return {'error': 'Nema selektovanih artikala'}, 400

    if is_active is None:
        return {'error': 'is_active je obavezan'}, 400

    updated = SupplierListing.query.filter(
        SupplierListing.id.in_(listing_ids),
        SupplierListing.supplier_id == g.supplier_id
    ).update({
        'is_active': bool(is_active),
        'updated_at': datetime.utcnow()
    }, synchronize_session='fetch')

    db.session.commit()

    action = 'aktivirano' if is_active else 'deaktivirano'
    return {
        'success': True,
        'message': f'{action.capitalize()} {updated} artikala',
        'updated': updated
    }


# ============== Stats ==============

@bp.route('/stats', methods=['GET'])
@supplier_jwt_required
def get_stats():
    """Get listing statistics"""
    total = SupplierListing.query.filter_by(supplier_id=g.supplier_id).count()
    active = SupplierListing.query.filter_by(supplier_id=g.supplier_id, is_active=True).count()
    in_stock = SupplierListing.query.filter(
        SupplierListing.supplier_id == g.supplier_id,
        SupplierListing.is_active == True,
        SupplierListing.stock_quantity > 0
    ).count()
    out_of_stock = active - in_stock

    categories = db.session.query(
        SupplierListing.part_category,
        db.func.count(SupplierListing.id)
    ).filter_by(
        supplier_id=g.supplier_id,
        is_active=True
    ).group_by(SupplierListing.part_category).all()

    return {
        'total_listings': total,
        'active_listings': active,
        'in_stock': in_stock,
        'out_of_stock': out_of_stock,
        'categories': {cat or 'uncategorized': count for cat, count in categories}
    }


# ============== Brands ==============

@bp.route('/brands', methods=['GET'])
@supplier_jwt_required
def list_brands():
    """Return list of standard brands for dropdown"""
    return {'brands': get_brand_list()}


# ============== JSON Import (Legacy) ==============

@bp.route('/import', methods=['POST'])
@supplier_jwt_required
def import_listings():
    """Import listings from JSON array"""
    data = request.json or {}
    listings_data = data.get('listings', [])

    if not listings_data:
        return {'error': 'No listings provided'}, 400

    eur_rate = get_supplier_eur_rate()
    created = 0
    updated = 0
    errors = []

    for item in listings_data:
        try:
            part_number = item.get('part_number')
            existing = None

            if part_number:
                existing = SupplierListing.query.filter_by(
                    supplier_id=g.supplier_id,
                    part_number=part_number
                ).first()

            # Calculate prices
            p_rsd = item.get('price_rsd') or item.get('price')
            p_eur = item.get('price_eur')
            prices = calculate_dual_prices(price_rsd=p_rsd, price_eur=p_eur, eur_rate=eur_rate)

            if existing:
                existing.name = item.get('name', existing.name)
                existing.brand = item.get('brand', existing.brand)
                existing.price_rsd = prices['price_rsd']
                existing.price_eur = prices['price_eur']
                existing.price = prices['price_rsd']
                existing.stock_quantity = item.get('stock_quantity', existing.stock_quantity)
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                listing = SupplierListing(
                    supplier_id=g.supplier_id,
                    name=item.get('name', 'Untitled'),
                    brand=item.get('brand'),
                    model_compatibility=item.get('model_compatibility'),
                    part_category=item.get('part_category'),
                    part_number=part_number,
                    description=item.get('description'),
                    is_original=item.get('is_original', False),
                    quality_grade=item.get('quality_grade'),
                    price_rsd=prices['price_rsd'],
                    price_eur=prices['price_eur'],
                    price=prices['price_rsd'],
                    currency='RSD',
                    min_order_qty=item.get('min_order_qty', 1),
                    stock_quantity=item.get('stock_quantity', 0),
                    delivery_days=item.get('delivery_days'),
                    is_active=True
                )
                db.session.add(listing)
                created += 1

        except Exception as e:
            errors.append(f"Error importing {item.get('name', 'unknown')}: {str(e)}")

    db.session.commit()

    return {
        'success': True,
        'data': {
            'created': created,
            'updated': updated,
            'errors': errors if errors else None
        }
    }


# ============== Excel Import ==============

@bp.route('/import-excel', methods=['POST'])
@supplier_jwt_required
def import_excel():
    """Import listings from XLSX/XLS file"""
    # 1. Proveri content length
    if request.content_length and request.content_length > MAX_IMPORT_FILE_SIZE_MB * 1024 * 1024:
        return {'error': f'Fajl je prevelik (max {MAX_IMPORT_FILE_SIZE_MB}MB)'}, 413

    # 2. Proveri da fajl postoji
    if 'file' not in request.files:
        return {'error': 'Fajl nije prilozen'}, 400

    file = request.files['file']
    if not file.filename:
        return {'error': 'Fajl nije izabran'}, 400

    # 3. Procitaj sadrzaj
    file_content = file.read()
    if not file_content:
        return {'error': 'Fajl je prazan'}, 400

    # 4. Security validacija
    is_valid, error_msg, safe_filename = validate_upload(
        file_content=file_content,
        filename=file.filename,
        allowed_extensions=['xlsx', 'xls'],
        max_size_mb=MAX_IMPORT_FILE_SIZE_MB,
        check_executable=True,
        check_office_macros=True
    )

    if not is_valid:
        return {'error': error_msg}, 400

    # 5. Parse Excel
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True, read_only=True)
        ws = wb.active

        if ws is None:
            return {'error': 'Excel fajl nema radni list'}, 400

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return {'error': 'Excel fajl nema podatke (samo header ili prazan)'}, 400

        if len(rows) - 1 > MAX_IMPORT_ROWS:
            return {'error': f'Previse redova ({len(rows) - 1}). Max: {MAX_IMPORT_ROWS}'}, 413

    except Exception as e:
        logger.error(f"Excel parse error: {e}")
        return {'error': f'Greska pri citanju Excel fajla: {str(e)}'}, 400

    # 6. Mapiraj header kolone
    raw_headers = rows[0]
    col_mapping = {}  # index -> field name

    for idx, header in enumerate(raw_headers):
        if header is None:
            continue
        h = str(header).strip().lower().replace(' ', '_')
        if h in COLUMN_MAP:
            col_mapping[idx] = COLUMN_MAP[h]

    # Proveri obavezne kolone
    mapped_fields = set(col_mapping.values())
    if 'name' not in mapped_fields:
        return {'error': 'Nedostaje obavezna kolona: naziv/name'}, 400
    if 'price_rsd' not in mapped_fields and 'price_eur' not in mapped_fields:
        return {'error': 'Nedostaje bar jedna kolona za cenu: cena_rsd/price_rsd ili cena_eur/price_eur'}, 400

    # 7. Ucitaj supplier EUR rate
    eur_rate = get_supplier_eur_rate()

    # 8. Iteriraj redove
    created = 0
    updated = 0
    skipped = 0
    errors = []
    data_rows = rows[1:]

    for row_num, row in enumerate(data_rows, start=2):
        try:
            # Mapiraj red u dict
            row_data = {}
            for idx, field in col_mapping.items():
                if idx < len(row):
                    val = row[idx]
                    if val is not None:
                        row_data[field] = val

            # Skip prazne redove
            name = str(row_data.get('name', '')).strip() if row_data.get('name') else ''
            if not name:
                skipped += 1
                continue

            # Parsiraj cene
            raw_rsd = row_data.get('price_rsd')
            raw_eur = row_data.get('price_eur')
            p_rsd = None
            p_eur = None

            if raw_rsd is not None:
                try:
                    p_rsd = float(str(raw_rsd).replace(',', '.').strip())
                    if p_rsd <= 0:
                        p_rsd = None
                except (ValueError, TypeError):
                    pass

            if raw_eur is not None:
                try:
                    p_eur = float(str(raw_eur).replace(',', '.').strip())
                    if p_eur <= 0:
                        p_eur = None
                except (ValueError, TypeError):
                    pass

            if p_rsd is None and p_eur is None:
                errors.append({'row': row_num, 'error': 'Nema validne cene', 'data': {'name': name}})
                continue

            prices = calculate_dual_prices(price_rsd=p_rsd, price_eur=p_eur, eur_rate=eur_rate)

            # Parsiraj ostale vrednosti
            part_number = str(row_data.get('part_number', '')).strip() if row_data.get('part_number') else None
            brand = str(row_data.get('brand', '')).strip() if row_data.get('brand') else None
            part_category = str(row_data.get('part_category', '')).strip() if row_data.get('part_category') else None
            quality_grade = str(row_data.get('quality_grade', '')).strip() if row_data.get('quality_grade') else None
            model_compat = str(row_data.get('model_compatibility', '')).strip() if row_data.get('model_compatibility') else None
            description = str(row_data.get('description', '')).strip() if row_data.get('description') else None

            stock_qty = 0
            if row_data.get('stock_quantity') is not None:
                try:
                    stock_qty = max(0, int(float(str(row_data['stock_quantity']))))
                except (ValueError, TypeError):
                    pass

            delivery_days = None
            if row_data.get('delivery_days') is not None:
                try:
                    delivery_days = max(0, int(float(str(row_data['delivery_days']))))
                except (ValueError, TypeError):
                    pass

            # Upsert po part_number
            existing = None
            if part_number:
                existing = SupplierListing.query.filter_by(
                    supplier_id=g.supplier_id,
                    part_number=part_number
                ).first()

            if existing:
                existing.name = name
                if brand:
                    existing.brand = brand
                if part_category:
                    existing.part_category = part_category
                if quality_grade:
                    existing.quality_grade = quality_grade
                if model_compat:
                    existing.model_compatibility = model_compat
                if description:
                    existing.description = description
                existing.price_rsd = prices['price_rsd']
                existing.price_eur = prices['price_eur']
                existing.price = prices['price_rsd']
                existing.stock_quantity = stock_qty
                if delivery_days is not None:
                    existing.delivery_days = delivery_days
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                listing = SupplierListing(
                    supplier_id=g.supplier_id,
                    name=name,
                    brand=brand,
                    model_compatibility=model_compat,
                    part_category=part_category,
                    part_number=part_number,
                    description=description,
                    quality_grade=quality_grade,
                    price_rsd=prices['price_rsd'],
                    price_eur=prices['price_eur'],
                    price=prices['price_rsd'],
                    currency='RSD',
                    stock_quantity=stock_qty,
                    delivery_days=delivery_days,
                    is_active=True
                )
                db.session.add(listing)
                created += 1

        except Exception as e:
            errors.append({'row': row_num, 'error': str(e), 'data': {'name': name if 'name' in dir() else '?'}})

    db.session.commit()

    return {
        'success': True,
        'data': {
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'errors': errors if errors else None
        }
    }


# ============== LCD Ponuda Import ==============

@bp.route('/import-lcd-ponuda', methods=['POST'])
@supplier_jwt_required
def import_lcd_ponuda():
    """Import LCD ponuda format XLS with optional preview.

    Query params:
        preview=true (default) - parse and return summary without DB insert
        preview=false - delete existing listings and import all
    """
    from app.services.lcd_ponuda_parser import parse_lcd_ponuda

    preview = request.args.get('preview', 'true').lower() != 'false'

    # 1. Validate file
    if 'file' not in request.files:
        return {'error': 'Fajl nije prilozen'}, 400

    file = request.files['file']
    if not file.filename:
        return {'error': 'Fajl nije izabran'}, 400

    file_content = file.read()
    if not file_content:
        return {'error': 'Fajl je prazan'}, 400

    if len(file_content) > MAX_IMPORT_FILE_SIZE_MB * 1024 * 1024:
        return {'error': f'Fajl je prevelik (max {MAX_IMPORT_FILE_SIZE_MB}MB)'}, 413

    # 2. Security validation
    is_valid, error_msg, safe_filename = validate_upload(
        file_content=file_content,
        filename=file.filename,
        allowed_extensions=['xlsx', 'xls'],
        max_size_mb=MAX_IMPORT_FILE_SIZE_MB,
        check_executable=True,
        check_office_macros=True
    )
    if not is_valid:
        return {'error': error_msg}, 400

    # 3. Get supplier EUR rate
    eur_rate = get_supplier_eur_rate()

    # 4. Parse
    try:
        result = parse_lcd_ponuda(file_content, eur_rate=eur_rate)
    except Exception as e:
        logger.error(f"LCD ponuda parse error: {e}")
        return {'error': f'Greska pri parsiranju fajla: {str(e)}'}, 400

    if not result['listings']:
        return {'error': 'Nisu pronadjeni artikli u fajlu'}, 400

    # 5. Preview mode - return summary only
    if preview:
        return {
            'success': True,
            'preview': True,
            'data': {
                **result['summary'],
                'sample': result['sample'],
            }
        }

    # 6. Import mode - delete existing and insert all
    existing = SupplierListing.query.filter_by(supplier_id=g.supplier_id).count()
    SupplierListing.query.filter_by(supplier_id=g.supplier_id).delete()
    db.session.flush()

    created = 0
    for item in result['listings']:
        listing = SupplierListing(
            supplier_id=g.supplier_id,
            name=item['name'],
            brand=item['brand'],
            model_compatibility=item['model_compatibility'],
            part_category=item['part_category'],
            part_number=item['part_number'],
            quality_grade=item['quality_grade'],
            is_original=item['is_original'],
            price=Decimal(str(item['price_eur'])),  # legacy NOT NULL
            price_eur=Decimal(str(item['price_eur'])),
            price_rsd=Decimal(str(item['price_rsd'])),
            stock_status=item['stock_status'],
            is_active=item['is_active'],
            description=item['description'],
            min_order_qty=item['min_order_qty'],
            currency=item['currency'],
        )
        db.session.add(listing)
        created += 1

    db.session.commit()

    return {
        'success': True,
        'preview': False,
        'data': {
            'deleted': existing,
            'created': created,
            'skipped': result['summary']['skipped'],
        }
    }


# ============== Excel Export ==============

@bp.route('/export-excel', methods=['GET'])
@supplier_jwt_required
def export_excel():
    """Export all listings as XLSX file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    listings = SupplierListing.query.filter_by(
        supplier_id=g.supplier_id
    ).order_by(SupplierListing.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cenovnik'

    # Header (matching template column order)
    headers = ['Naziv', 'Sifra', 'Brend', 'Kategorija', 'Cena EUR', 'Cena RSD',
               'Kolicina', 'Kvalitet', 'Modeli', 'Rok isporuke', 'Opis', 'Aktivan']
    ws.append(headers)

    # Bold header with purple fill
    header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for l in listings:
        ws.append([
            l.name,
            l.part_number,
            l.brand,
            l.part_category,
            float(l.price_eur) if l.price_eur else None,
            float(l.price_rsd) if l.price_rsd else None,
            l.stock_quantity,
            l.quality_grade,
            l.model_compatibility,
            l.delivery_days,
            l.description,
            'Da' if l.is_active else 'Ne'
        ])

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-width kolone
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    supplier = Supplier.query.get(g.supplier_id)
    slug = supplier.slug if supplier else 'supplier'
    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    filename = f'cenovnik_{slug}_{date_str}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============== Excel Template ==============

@bp.route('/template', methods=['GET'])
@supplier_jwt_required
def download_template():
    """Download Excel template with headers, examples, and instructions"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ===== Sheet 1: Cenovnik (data entry) =====
    ws = wb.active
    ws.title = 'Cenovnik'

    headers = ['Naziv', 'Sifra', 'Brend', 'Kategorija', 'Cena EUR', 'Cena RSD',
               'Kolicina', 'Kvalitet', 'Modeli', 'Rok isporuke', 'Opis']
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(
        bottom=Side(style='thin', color='4B0082')
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Example rows - showing variant pattern (same model, different quality/price)
    example_font = Font(italic=True, color='888888', size=10)
    examples = [
        ['Display iPhone 15 Pro - Service Pack', 'A-IP15P-SP', 'Apple', 'display', 180, None, 5, 'service_pack', 'iPhone 15 Pro, 15 Pro Max', 1, 'Originalni Service Pack sa frejmom'],
        ['Display iPhone 15 Pro - OLED Hard', 'A-IP15P-OH', 'Apple', 'display', 95, None, 10, 'oled_hard', 'iPhone 15 Pro, 15 Pro Max', 1, 'OLED Hard zamena'],
        ['Display iPhone 15 Pro - OLED Soft', 'A-IP15P-OS', 'Apple', 'display', 75, None, 15, 'oled_soft', 'iPhone 15 Pro, 15 Pro Max', 1, 'OLED Soft zamena'],
        ['Display iPhone 15 Pro - TFT Incell', 'A-IP15P-TFT', 'Apple', 'display', 38, None, 20, 'tft_incell', 'iPhone 15 Pro, 15 Pro Max', 2, 'TFT Incell zamena'],
        ['Display Samsung S24 Ultra - Service Pack', 'S-S24U-SP', 'Samsung', 'display', 320, None, 3, 'service_pack', 'Galaxy S24 Ultra', 1, ''],
        ['Display Samsung S24 Ultra - OLED', 'S-S24U-OLED', 'Samsung', 'display', 140, None, 8, 'oled_hard', 'Galaxy S24 Ultra', 1, ''],
        ['Baterija iPhone 15', 'A-IP15-BAT', 'Apple', 'battery', 12, None, 30, 'oem', 'iPhone 15', 1, ''],
        ['Port punjenja Samsung S23', 'S-S23-CHG', 'Samsung', 'charging', 8, None, 50, '', 'Galaxy S23, S23+', 2, ''],
    ]
    for row_data in examples:
        ws.append(row_data)

    for row_num in range(2, len(examples) + 2):
        for cell in ws[row_num]:
            cell.font = example_font

    # Auto-width columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # ===== Sheet 2: Uputstvo =====
    ws2 = wb.create_sheet('Uputstvo')
    instructions = [
        ['UPUTSTVO ZA POPUNJAVANJE CENOVNIKA'],
        [''],
        ['KOLONE:'],
        ['Naziv', 'Obavezno. Ime artikla (npr. "Display iPhone 15 Pro - OLED Hard")'],
        ['Sifra', 'Preporuceno. Vasa interna sifra. Koristi se za azuriranje postojecih artikala.'],
        ['Brend', 'Preporuceno. Apple, Samsung, Huawei, Xiaomi, Motorola, Google, Oppo, Vivo...'],
        ['Kategorija', 'display, battery, camera, charging, speaker, frame, back_cover, flex, connector, ic, touch, other'],
        ['Cena EUR', 'Unesite bar jednu cenu (EUR ili RSD). Sistem preracunava drugu automatski.'],
        ['Cena RSD', 'Ako unesete obe cene, EUR ima prioritet.'],
        ['Kolicina', 'Broj komada na stanju. Ostavite prazno za neograniceno.'],
        ['Kvalitet', 'service_pack, original, oled_hard, oled_soft, oem, tft_incell, aaa, copy'],
        ['Modeli', 'Kompatibilni modeli, razdvojeni zarezom (npr. "iPhone 15, 15 Pro, 15 Pro Max")'],
        ['Rok isporuke', 'Broj dana za isporuku'],
        ['Opis', 'Opcioni opis artikla'],
        [''],
        ['PRAVILA:'],
        ['- Redovi sa praznim nazivom se preskacu'],
        ['- Ako artikal sa istom sifrom vec postoji, azurira se (upsert)'],
        ['- Primer redovi (italik) su samo za demonstraciju - obrisite ih pre uvoza'],
        ['- Max 5000 redova po fajlu, max 5MB velicina'],
        ['- EUR cene se automatski preracunavaju po vasem kursu'],
        [''],
        ['VARIJANTE KVALITETA (za displeje):'],
        ['service_pack', 'Originalni deo sa frejmom, direktno od proizvodjaca'],
        ['original', 'Originalni deo (izvadjen/refurbished)'],
        ['oled_hard', 'Aftermarket OLED - hard tip (visi kvalitet)'],
        ['oled_soft', 'Aftermarket OLED - soft/flex tip'],
        ['oem', 'OEM kvalitet'],
        ['tft_incell', 'TFT Incell zamena (budget opcija)'],
        ['aaa', 'AAA kopija'],
        ['copy', 'Obicna kopija'],
    ]

    title_font = Font(bold=True, size=14, color='4B0082')
    section_font = Font(bold=True, size=11, color='333333')
    normal_font = Font(size=10, color='555555')
    key_font = Font(bold=True, size=10, color='333333')

    for row_data in instructions:
        ws2.append(row_data)

    # Style instructions
    ws2['A1'].font = title_font
    ws2['A3'].font = section_font
    ws2['A16'].font = section_font
    ws2['A22'].font = section_font

    for row_num in range(4, 15):
        ws2.cell(row=row_num, column=1).font = key_font
        ws2.cell(row=row_num, column=2).font = normal_font
    for row_num in range(23, 31):
        ws2.cell(row=row_num, column=1).font = key_font
        ws2.cell(row=row_num, column=2).font = normal_font

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 70

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='sablon_cenovnik.xlsx'
    )
