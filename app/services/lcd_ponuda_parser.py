"""
LCD Ponuda XLS Parser Service.

Parses section-based LCD price list XLS files from suppliers.
Format: sections by brand with header rows (A="SIFRA"), data rows below.

Columns: A=part_number, B=model, C=color, D=price_EUR, E=quality_description
"""
import re
from decimal import Decimal
from io import BytesIO
import openpyxl


# Known manufacturer brands found in quality descriptions (E column)
KNOWN_MFR_BRANDS = ['JK', 'GX', 'DD', 'ZY', 'RJ', 'TC', 'YK', 'HX', 'ZJ', 'MP', 'OG']

# Quality pattern matching (order matters - first match wins)
QUALITY_PATTERNS = [
    ('service pack', 'service_pack'),
    ('soft oled', 'oled_soft'),
    ('hard oled', 'oled_hard'),
    ('oled', 'oled_hard'),
    ('incell', 'tft_incell'),
    ('frejm', 'copy'),
]


def parse_lcd_ponuda(file_content, eur_rate=118.0):
    """Parse LCD ponuda XLS file.

    Args:
        file_content: bytes content of the XLS/XLSX file
        eur_rate: EUR to RSD exchange rate

    Returns:
        dict with 'listings', 'summary', 'sample' keys
    """
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active

    rate = Decimal(str(eur_rate))
    brand = None
    sec_quality = None
    sec_header = ''
    is_tablet = False
    listings = []
    skipped = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a = row[0].value
        b = row[1].value
        c = row[2].value
        d = row[3].value
        e = row[4].value if len(row) > 4 else None
        f_val = row[5].value if len(row) > 5 else None

        if all(cell.value is None for cell in row[:6]):
            continue

        # Section header row
        if a == 'SIFRA':
            sec_header_raw = str(b or '').strip()
            brand = _extract_brand(sec_header_raw)
            sec_quality = _extract_section_quality(sec_header_raw, str(c or ''), str(e or ''))
            is_tablet = 'TABLET' in sec_header_raw.upper()
            sec_header = f"{sec_header_raw} {c or ''} {d or ''} {e or ''} {f_val or ''}"
            continue

        # Skip rows without brand context or price
        if brand is None or d is None or not isinstance(d, (int, float)):
            skipped += 1
            continue

        pn = str(a).strip() if a else None
        model = _clean_model(b, brand)
        if not model:
            skipped += 1
            continue

        desc = str(e).strip() if e else None
        row_q, mfr_brand = _normalize_quality_with_mfr(desc)
        qg = row_q or sec_quality or 'copy'

        # For service_pack, default mfr brand is IQKO
        if qg == 'service_pack' and not mfr_brand:
            mfr_brand = 'IQKO'

        price_eur = Decimal(str(d))
        price_rsd = (price_eur * rate).quantize(Decimal('0.01'))
        chip = _has_chip(b) if brand == 'Apple' else None
        frame = _get_frame(sec_header, c, desc)

        # Extract color from column C (e.g. BLACK, WHITE, GREEN, NO FRAME)
        raw_color = str(c).strip().upper() if c else None
        color = None
        if raw_color and raw_color != 'NONE' and raw_color != 'NO FRAME':
            color = raw_color

        name = _build_name(brand, model, qg, frame, chip, is_tablet, mfr_brand, color)

        listings.append({
            'name': name[:200],
            'brand': brand[:50],
            'model_compatibility': model,
            'part_category': 'display',
            'part_number': pn[:50] if pn else None,
            'quality_grade': qg,
            'is_original': qg == 'service_pack',
            'price_eur': float(price_eur),
            'price_rsd': float(price_rsd),
            'stock_status': 'IN_STOCK',
            'is_active': True,
            'description': desc,
            'color': color,
            'min_order_qty': 1,
            'currency': 'EUR',
            'mfr_brand': mfr_brand,
        })

    wb.close()

    # Build summary
    by_brand = {}
    by_quality = {}
    prices = []
    for l in listings:
        by_brand[l['brand']] = by_brand.get(l['brand'], 0) + 1
        by_quality[l['quality_grade']] = by_quality.get(l['quality_grade'], 0) + 1
        prices.append(l['price_eur'])

    summary = {
        'total': len(listings),
        'skipped': skipped,
        'by_brand': dict(sorted(by_brand.items(), key=lambda x: -x[1])),
        'by_quality': dict(sorted(by_quality.items(), key=lambda x: -x[1])),
        'price_range': {
            'min_eur': min(prices) if prices else 0,
            'max_eur': max(prices) if prices else 0,
        },
    }

    sample = []
    for l in listings[:20]:
        sample.append({
            'name': l['name'],
            'brand': l['brand'],
            'model': l['model_compatibility'],
            'price_eur': l['price_eur'],
            'price_rsd': l['price_rsd'],
            'quality_grade': l['quality_grade'],
            'mfr_brand': l['mfr_brand'],
        })

    return {
        'listings': listings,
        'summary': summary,
        'sample': sample,
    }


# ============== Helper Functions ==============

def _extract_brand(header):
    """Extract phone brand from section header."""
    h = header.upper()
    for name in ['GOOGLE', 'APPLE', 'SAMSUNG', 'HUAWEI', 'HONOR', 'XIAOMI',
                 'REALME', 'MOTOROLA', 'OPPO', 'VIVO']:
        if name in h:
            return 'Huawei' if name == 'HONOR' else name.title()
    return h.split()[0].title() if h.strip() else 'Unknown'


def _extract_section_quality(header, grade, quality):
    """Extract quality grade from section header."""
    full = f"{header} {grade or ''} {quality or ''}".upper()
    if 'SERVICE' in full and 'PACK' in full:
        return 'service_pack'
    if 'SOFT OLED' in full and 'HARD' not in full:
        return 'oled_soft'
    if 'HARD OLED' in full and 'SOFT' not in full:
        return 'oled_hard'
    if 'HARD' in full and ('SOFT' in full or 'OLED' in full):
        return 'oled_hard'
    if 'TFT' in full or 'INCELL' in full:
        return 'tft_incell'
    if 'BEST' in full and 'COPY' in full:
        return 'copy'
    if 'FULL ORG' in full:
        return 'oem'
    return None


def _normalize_quality_with_mfr(text):
    """Extract quality grade AND manufacturer brand from E column.

    Returns:
        tuple: (quality_grade, mfr_brand) e.g. ('oled_soft', 'JK')
    """
    if not text:
        return None, None

    t = re.sub(r'\s+', ' ', str(text).strip().lower())
    t_upper = str(text).strip().upper()

    # Extract manufacturer brand prefix
    mfr_brand = None
    for mb in KNOWN_MFR_BRANDS:
        if t_upper.startswith(mb + ' ') or t_upper.startswith(mb + '  '):
            mfr_brand = mb
            break

    # Match quality pattern
    for pattern, grade in QUALITY_PATTERNS:
        if pattern in t:
            return grade, mfr_brand

    return None, mfr_brand


def _clean_model(raw, brand):
    """Clean model name from raw XLS data."""
    if raw is None:
        return None
    m = str(raw).strip()
    m = re.sub(r'\(\s*[Bb]ez\s+[Cc]ipa\s*\)', '', m)
    m = re.sub(r'\(\s*[Ss][Aa]\s+[Cc][Ii][Pp][Oo][Mm]\s*\)', '', m)
    m = re.sub(r'\(TFT\s*/\s*incell\)', '', m)
    m = re.sub(r'\s*(SOFT|HARD)\s+OLED\s*', '', m, flags=re.IGNORECASE)
    m = re.sub(r'\s+', ' ', m).strip()
    if brand == 'Apple' and m and re.match(r'^(\d|X|SE)', m, re.IGNORECASE):
        m = 'iPhone ' + m
    if brand == 'Google' and m and not m.lower().startswith('pixel'):
        m = 'Pixel ' + m
    return m if m else None


def _has_chip(raw):
    """Check if Apple model includes chip info."""
    if not raw:
        return None
    u = str(raw).upper()
    if 'SA CIPOM' in u:
        return True
    if 'BEZ CIPA' in u:
        return False
    return None


def _get_frame(header_full, color, desc):
    """Determine frame inclusion from header/color/description."""
    h = header_full.upper()
    if 'SA FREJMOM' in h or ('FREJMOM' in h and 'BEZ' not in h):
        return True
    if 'BEZ FREJMA' in h or ('FREJMA' in h and 'SA' not in h):
        return False
    if desc:
        d = desc.lower()
        if 'sa frejmom' in d:
            return True
        if 'bez frejm' in d:
            return False
    if color and 'NO FRAME' in str(color).upper():
        return False
    return None


def _build_name(brand, model, qg, frame, chip, tablet, mfr_brand=None, color=None):
    """Build listing name from components."""
    parts = ['Display Tablet' if tablet else 'Display']
    if brand:
        parts.append(brand)
    if model:
        parts.append(model)

    # Color (e.g., BLACK, GREEN, CREAM)
    if color:
        parts.append(color.title())

    labels = {
        'service_pack': 'SP',
        'oled_hard': 'Hard OLED',
        'oled_soft': 'Soft OLED',
        'tft_incell': 'TFT Incell',
        'oem': 'OEM',
        'copy': 'Copy',
    }
    if qg in labels:
        parts.append(f'({labels[qg]})')

    # Manufacturer brand (e.g., IQKO, JK, GX, YK)
    if mfr_brand:
        parts.append(mfr_brand)

    if frame is True:
        parts.append('sa frejmom')
    elif frame is False and qg not in ('service_pack',):
        parts.append('bez frejma')
    if chip is True:
        parts.append('sa cipom')
    elif chip is False and qg == 'service_pack':
        parts.append('bez cipa')

    return ' '.join(parts)
